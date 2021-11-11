import openpyxl
import click
import os
from copy import copy

def copyRange(trans,sheet):
    rangeSelected = []
    for row in sheet.iter_rows():
        rowSelected = []
        for cell in row:
            if trans == 'caps':
                val = cell.value.upper()
            else:
                val = cell.value
            new_cell = sheet.cell(row=cell.row, column=cell.col_idx,
                value= val)
            if cell.has_style and trans == 'pres':
                new_cell._style = copy(cell.style)
            rowSelected.append(cell.value)
        rangeSelected.append(rowSelected)
    return rangeSelected

def pasteRange(sheetReceiving, copiedData):
    for row in copiedData:
        sheetReceiving.append(row)

@click.command(help="copy the source file to destination file")
@click.option('--capitalize','transformation',flag_value='caps',default=False,help='capitalize the content in destination')
@click.option('--preservestyles','transformation',flag_value='pres',default=False,help='capitalize the content in destination')
@click.argument('files',nargs=2)
def copy_files(transformation,files):
    input =files[0]
    output=files[1]

    src = openpyxl.load_workbook(input)
    sheetnames = src.get_sheet_names()
    sheet_to_copy = src.get_sheet_by_name(sheetnames[1])


    try :
        if (os.path.exists(output) == 1):
            res = click.prompt("Do you want to overwrite?")
            if res == 'yes' or res == 'Yes' or res == 'YES':
                dest = openpyxl.load_workbook(output)
    finally:
        dest = openpyxl.Workbook()

    dest.create_sheet('Sheet1')
    sheet = dest.get_sheet_by_name('Sheet1')

    selectedRange = copyRange( transformation,sheet_to_copy)
    pasteRange(sheet, selectedRange)
    dest.save(output)


if __name__=='__main__':
    copy_files()