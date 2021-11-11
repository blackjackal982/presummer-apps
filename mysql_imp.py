import click
import mysql.connector
import openpyxl
from mysql.connector import Error
def get_data(sheet):
    rangeSelected = []
    for row in sheet.iter_rows():
        rowSelected = []
        for cell in row:
            rowSelected.append(cell.value)
        rangeSelected.append(rowSelected)
    return rangeSelected[1:]

def modify_data(data):
    rangeSelected = []
    for row in data:
        rowSelected = []
        rowSelected.append("ol2016_"+row[1].lower()+"_"+row[3].lower()+"_mock")
        rowSelected.append(row[1])
        rowSelected.append(row[2])
        rangeSelected.append(rowSelected)
    return rangeSelected

@click.group()
def main():
    click.echo("main")

@main.command(help="create database")
@click.argument("dbname")
def createdb(dbname):
    try:
        conn = mysql.connector.connect(host="localhost", user="root", password="dingoo12")
        print(conn)
        mycursor = conn.cursor()
        mycursor.execute("CREATE DATABASE "+dbname+";")
        click.echo("database created successfully")
    except Error as e:
        click.echo(e)
    finally:
        mycursor.close()
        conn.close()


@main.command(help="drop database")
@click.argument("dbname")
def dropdb(dbname):
    try:
        conn = mysql.connector.connect(host="localhost", user="root", password="dingoo12")
        mycursor = conn.cursor()
        mycursor.execute("DROP DATABASE " + dbname+";")
        click.echo("database dropped succesfully")
    except Error as e:
        click.echo(e)
    finally:
        mycursor.close()
        conn.close()

@main.command(help="import the data into tables")
@click.argument("dbname")
@click.argument("file_1")
@click.argument("file_2")
def importdata(dbname,file_1,file_2):
    try:
        conn = mysql.connector.connect(host="localhost", user="root", password="dingoo12",database=dbname)
        mycursor = conn.cursor()
        print(conn)
        #tables created
        mycursor.execute("CREATE TABLE IF NOT EXISTS student\
                (NAME VARCHAR(100),\
                CLG_NAME VARCHAR(100) NOT NULL,\
                EMAIL_ID VARCHAR(100) NOT NULL,\
                         PRIMARY KEY(NAME));")
        mycursor.execute("CREATE TABLE IF NOT EXISTS marks\
                         (NAME VARCHAR(100),\
                           TOTAL_MARKS INT,PRIMARY KEY(NAME),\
                            FOREIGN KEY(NAME) REFERENCES student(NAME));")

        #read data from student.xlsx and result.xlsx
        workbook_1 = openpyxl.load_workbook(file_1)
        sheet_obj_1= workbook_1.get_sheet_by_name("Current")
        student_data = get_data(sheet_obj_1)
        student_data = modify_data(student_data)


        workbook_2 = openpyxl.load_workbook(file_2)
        sheet_obj_2= workbook_2.get_sheet_by_name("Mock_Res")
        marks_data = get_data(sheet_obj_2)
        del marks_data[96] #entry for which student data doesn't exist used bcz excel is not working so deletingwrong entry here
        for i in range(len(marks_data)):
            marks_data[i]= [marks_data[i][0],int(marks_data[i][-1])]

        #store the data into tables
        sql_query = " INSERT INTO student(NAME,CLG_NAME,EMAIL_ID) VALUES (%s,%s,%s)"
        mycursor.executemany(sql_query,student_data)
        conn.commit()

        sql_query = " INSERT INTO marks(NAME,TOTAL_MARKS) VALUES (%s,%s)"
        mycursor.executemany(sql_query, marks_data)
        conn.commit()

    except Error as e:
        click.echo(e)
    finally:
        mycursor.close()
        conn.close()

@main.command(help="prints college status on console")
@click.argument("dbname")
@click.argument('table1')
@click.argument('table2')
def collegestats(dbname,table1,table2):
    try:
        conn = mysql.connector.connect(host="localhost", user="root", password="dingoo12",database =dbname)
        mycursor = conn.cursor()
        query = "SELECT clg_name,count(*),max(total_marks),min(total_marks),avg(total_marks) \
        from "+table1+" join "+table2+" on student.name=marks.name group by clg_name;"
        mycursor.execute(query)
        records = mycursor.fetchall()
        click.echo("college\tno_of_students\tmax_marks\tmin_marks\tavg_marks")
        for row in records:
            click.echo( row[0]+"\t\t"+str(row[1])+"\t\t"+str(row[2])+"\t\t"+str(row[3])+"\t\t"+str(row[4])+"\n")
    except Error as e:
        click.echo(e)
    finally:
        mycursor.close()
        conn.close()

if __name__ == '__main__':
    main()