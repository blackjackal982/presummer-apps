import click
import openpyxl
from bs4 import BeautifulSoup
import os

def pasteRange(sheetReceiving, copiedData):
    for row in copiedData:
        sheetReceiving.append(row)

@click.command(help="scrape html file and dump to excel")
@click.argument('input',type=click.File())
@click.argument('output')
def dump_to_excel(input,output):
    file = open(input.name)
    soup = BeautifulSoup(file,'html.parser')
    table_header =[i.get_text(strip=True) for i in soup.find_all('th')][1:]
    table_row = soup.find_all('tr')

    rows=[]
    rows.append(table_header)
    for i in range(1,len(table_row)):
        data = [i.get_text(strip=True) for i in table_row[i].find_all('td')][1:]
        rows.append(data)

    try:
        if (os.path.exists(output) == 1):
            dest = openpyxl.load_workbook(output)
    finally:
        dest = openpyxl.Workbook()

    dest.create_sheet('Mock_Res')
    sheet = dest.get_sheet_by_name('Mock_Res')
    pasteRange(sheet,rows)
    dest.save(output)

if __name__=='__main__':
    dump_to_excel()


